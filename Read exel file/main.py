# Import openyxl module
import openpyxl
import json


# Define variable to load the wookbook
wookbook = openpyxl.load_workbook("nums.xlsx")

# Define variable to read the active sheet:
worksheet = wookbook.active


list_numbers = dict()

# Iterate the loop to read the cell values
for i in range(0, worksheet.max_row):
    for col in worksheet.iter_cols(1, worksheet.max_column):
        print(col[i].value, end="\t\t")
        list_numbers[i] = col[i].value
    print('')

print(list_numbers)

with open('numbers.json', 'w') as f:
    json_format = json.dumps(list_numbers, ensure_ascii=False).encode('utf-8')
    f.write(json_format.decode())
